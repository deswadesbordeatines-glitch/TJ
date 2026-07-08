import os
import re
import subprocess
import threading
import traceback
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from queue import Empty, Queue
from tkinter import Tk, StringVar, filedialog, messagebox
from tkinter import ttk

import openpyxl
import pdfplumber
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


APP_TITLE = "采购发票识别工具"
DEFAULT_EXCEL_NAME = "采购发票台账.xlsx"
DETAIL_HEADERS = [
    "发票类型",
    "发票号码",
    "开票日期",
    "购买方名称",
    "销售方名称",
    "金额",
    "税额",
    "价税合计",
    "来源文件",
    "导入时间",
]


def documents_dir() -> Path:
    return Path.home() / "Documents"


def default_excel_path() -> Path:
    return documents_dir() / DEFAULT_EXCEL_NAME


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_pdf_text(value: str) -> str:
    value = value.replace("\u3000", " ").replace("\xa0", " ")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n{2,}", "\n", value)
    return value.strip()


def first_match(patterns, text: str, flags: int = 0) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return clean_text(match.group(1))
    return ""


def parse_decimal(value: str):
    text = str(value or "").replace(",", "").replace("¥", "").replace("￥", "").strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def normalize_money(value: str) -> str:
    amount = parse_decimal(value)
    return f"{amount:.2f}" if amount is not None else clean_text(value)


def normalize_date(value: str) -> str:
    if not value:
        return ""
    text = value.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-").replace(".", "-")
    parts = text.split("-")
    if len(parts) == 3:
        return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return text


def write_log(path: Path, message: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with path.open("a", encoding="utf-8") as file:
        file.write(f"[{timestamp}] {message}\n")


def extract_text_with_pymupdf(pdf_path: Path, debug_log: Path) -> str:
    try:
        import fitz
    except Exception as exc:
        write_log(debug_log, f"PyMuPDF 不可用：{exc}")
        return ""

    try:
        pages = []
        with fitz.open(pdf_path) as document:
            write_log(debug_log, f"PyMuPDF 打开成功：{pdf_path.name}，页数 {document.page_count}")
            for index, page in enumerate(document, start=1):
                page_text = page.get_text("text") or ""
                pages.append(page_text)
                write_log(debug_log, f"PyMuPDF 第 {index} 页文字长度：{len(page_text)}")
                write_log(debug_log, f"PyMuPDF 第 {index} 页文字：\n{page_text[:8000]}")
        return normalize_pdf_text("\n".join(pages))
    except Exception as exc:
        write_log(debug_log, f"PyMuPDF 提取失败：{exc}")
        return ""


def extract_text_with_pdfplumber(pdf_path: Path, debug_log: Path) -> str:
    try:
        pages = []
        with pdfplumber.open(pdf_path) as pdf:
            write_log(debug_log, f"pdfplumber 打开成功：{pdf_path.name}，页数 {len(pdf.pages)}")
            for index, page in enumerate(pdf.pages, start=1):
                page_text = page.extract_text(x_tolerance=1.5, y_tolerance=3) or ""
                pages.append(page_text)
                write_log(debug_log, f"pdfplumber 第 {index} 页文字长度：{len(page_text)}")
                write_log(debug_log, f"pdfplumber 第 {index} 页文字：\n{page_text[:8000]}")
        return normalize_pdf_text("\n".join(pages))
    except Exception as exc:
        write_log(debug_log, f"pdfplumber 提取失败：{exc}")
        return ""


def extract_pdf_text(pdf_path: Path, debug_log: Path) -> str:
    pymupdf_text = extract_text_with_pymupdf(pdf_path, debug_log)
    pdfplumber_text = extract_text_with_pdfplumber(pdf_path, debug_log)
    combined = normalize_pdf_text("\n".join(part for part in [pymupdf_text, pdfplumber_text] if part))
    write_log(debug_log, f"合并后文字长度：{len(combined)}")
    if not combined:
        raise ValueError("未读取到 PDF 原始文字。本离线基础版不包含 OCR。")
    return combined


def remove_party_noise(value: str) -> str:
    for label in ["统一社会信用代码", "纳税人识别号", "地址", "电话", "开户行", "账号"]:
        value = value.split(label)[0]
    return clean_text(value).strip(":：")


def extract_party_name(text: str, party_label: str) -> str:
    if party_label == "购买方":
        value = first_match(
            [
                r"购\s*名称[:：\s]*([^\n]+?)\s+销\s*名称[:：\s]*[^\n]+",
                r"购买方(?:信息)?[\s\S]{0,180}?名称[:：\s]*([^\n]+?)(?:\s+统一社会信用代码|\s+纳税人识别号|\n)",
                r"购买方名称[:：\s]*([^\n]+)",
            ],
            text,
        )
    else:
        value = first_match(
            [
                r"购\s*名称[:：\s]*[^\n]+?\s+销\s*名称[:：\s]*([^\n]+)",
                r"销售方(?:信息)?[\s\S]{0,180}?名称[:：\s]*([^\n]+?)(?:\s+统一社会信用代码|\s+纳税人识别号|\n)",
                r"销售方名称[:：\s]*([^\n]+)",
            ],
            text,
        )
    return remove_party_noise(value)


def extract_invoice_type(text: str) -> str:
    return first_match(
        [
            r"(电子发票[（(][^）)]+[）)])",
            r"(增值税专用发票)",
            r"(增值税普通发票)",
        ],
        text,
    )


def extract_amounts(text: str):
    amount = first_match(
        [
            r"合\s*计\s*[¥￥]\s*([0-9,]+\.[0-9]{2})\s*[¥￥]\s*[0-9,]+\.[0-9]{2}",
            r"金额[:：\s]*[¥￥]?\s*([0-9,]+\.[0-9]{2})",
        ],
        text,
    )
    tax = first_match(
        [
            r"合\s*计\s*[¥￥]\s*[0-9,]+\.[0-9]{2}\s*[¥￥]\s*([0-9,]+\.[0-9]{2})",
            r"税额[:：\s]*[¥￥]?\s*([0-9,]+\.[0-9]{2})",
        ],
        text,
    )
    total = first_match(
        [
            r"价税合计[（(]大写[）)][\s\S]{0,80}?[（(]小写[）)]\s*[¥￥]\s*([0-9,]+\.[0-9]{2})",
            r"[（(]小写[）)]\s*[¥￥]\s*([0-9,]+\.[0-9]{2})",
            r"价税合计[\s\S]{0,80}?[¥￥]\s*([0-9,]+\.[0-9]{2})",
        ],
        text,
    )
    return normalize_money(amount), normalize_money(tax), normalize_money(total)


def parse_invoice(pdf_path: Path, debug_log: Path) -> dict:
    write_log(debug_log, "=" * 80)
    write_log(debug_log, f"开始识别：{pdf_path}")
    text = extract_pdf_text(pdf_path, debug_log)
    amount, tax, total = extract_amounts(text)
    data = {
        "发票类型": extract_invoice_type(text),
        "发票号码": first_match([r"发票号码[:：\s]*([0-9]{8,30})"], text),
        "开票日期": normalize_date(
            first_match(
                [
                    r"开票日期[:：\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)",
                    r"开票日期[:：\s]*([0-9]{4}[-/.][0-9]{1,2}[-/.][0-9]{1,2})",
                ],
                text,
            )
        ),
        "购买方名称": extract_party_name(text, "购买方"),
        "销售方名称": extract_party_name(text, "销售方"),
        "金额": amount,
        "税额": tax,
        "价税合计": total,
        "来源文件": pdf_path.name,
        "导入时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    missing = [name for name in ["发票号码", "开票日期", "购买方名称", "销售方名称", "金额", "税额", "价税合计"] if not data[name]]
    write_log(debug_log, "字段结果：" + "；".join(f"{key}={value or '<空>'}" for key, value in data.items()))
    if missing:
        raise ValueError("缺少关键字段：" + "、".join(missing))
    return data


def collect_pdf_files(paths) -> list:
    result = []
    seen = set()
    ignored_dirs = {".git", ".venv", ".build-venv", "__pycache__", "build", "dist", "node_modules"}
    for raw_path in paths:
        path = Path(raw_path)
        if path.is_file() and path.suffix.lower() == ".pdf":
            candidates = [path]
        elif path.is_dir():
            candidates = []
            for root, dir_names, file_names in os.walk(path):
                dir_names[:] = [
                    name for name in dir_names
                    if name not in ignored_dirs and not name.startswith(".") and not name.lower().endswith(".app")
                ]
                root_path = Path(root)
                for file_name in file_names:
                    candidate = root_path / file_name
                    if candidate.suffix.lower() == ".pdf":
                        candidates.append(candidate)
        else:
            candidates = []
        for candidate in sorted(candidates):
            resolved = str(candidate.resolve()).lower()
            if resolved not in seen:
                seen.add(resolved)
                result.append(candidate)
    return result


def ensure_workbook(excel_path: Path):
    if excel_path.exists():
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "发票台账"
        sheet.append(DETAIL_HEADERS)
    if sheet.max_row < 1 or sheet.cell(1, 1).value != DETAIL_HEADERS[0]:
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(DETAIL_HEADERS)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for index, header in enumerate(DETAIL_HEADERS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(14, min(28, len(header) * 2 + 6))
    return workbook, sheet


def write_excel(excel_path: Path, rows: list) -> None:
    workbook, sheet = ensure_workbook(excel_path)
    for row_data in rows:
        sheet.append([row_data.get(header, "") for header in DETAIL_HEADERS])
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)


class InvoiceApp:
    def __init__(self):
        self.root = Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("820x520")
        self.root.resizable(False, False)
        self.selected_paths = []
        self.output_path_var = StringVar(value=str(default_excel_path()))
        self.status_var = StringVar(value="请选择 PDF 文件或发票文件夹")
        self.count_var = StringVar(value="成功：0    失败：0")
        self.queue = Queue()
        self.worker = None
        self.build_ui()
        self.root.after(200, self.poll_queue)

    def build_ui(self):
        main = ttk.Frame(self.root, padding=18)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text=APP_TITLE, font=("Microsoft YaHei UI", 20, "bold")).pack(anchor="w")
        ttk.Label(main, text="Windows 离线版：仅本地识别电子发票 PDF 基础字段，不调用网络接口。").pack(anchor="w", pady=(6, 14))

        button_frame = ttk.Frame(main)
        button_frame.pack(fill="x", pady=(0, 12))
        ttk.Button(button_frame, text="选择 PDF 文件", command=self.choose_files).pack(side="left", padx=(0, 8))
        ttk.Button(button_frame, text="选择发票文件夹", command=self.choose_folder).pack(side="left", padx=(0, 8))
        ttk.Button(button_frame, text="开始识别", command=self.start).pack(side="left", padx=(0, 8))
        ttk.Button(button_frame, text="打开输出 Excel", command=self.open_excel).pack(side="left")

        output_frame = ttk.Frame(main)
        output_frame.pack(fill="x", pady=(0, 12))
        ttk.Label(output_frame, text="输出 Excel：").pack(side="left")
        ttk.Entry(output_frame, textvariable=self.output_path_var).pack(side="left", fill="x", expand=True, padx=(8, 8))
        ttk.Button(output_frame, text="选择输出位置", command=self.choose_output).pack(side="left")

        ttk.Label(main, textvariable=self.status_var).pack(anchor="w")
        ttk.Label(main, textvariable=self.count_var, font=("Microsoft YaHei UI", 12, "bold")).pack(anchor="w", pady=(8, 8))
        self.progress = ttk.Progressbar(main, maximum=100, mode="determinate")
        self.progress.pack(fill="x", pady=(0, 12))

        ttk.Label(main, text="日志").pack(anchor="w")
        log_frame = ttk.Frame(main)
        log_frame.pack(fill="both", expand=True)
        self.log_text = __import__("tkinter").Text(log_frame, height=12, wrap="word", state="disabled")
        scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scroll.set)
        self.log_text.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.log("程序已启动。")

    def log(self, message: str):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def choose_files(self):
        files = filedialog.askopenfilenames(title="选择 PDF 文件", filetypes=[("PDF 文件", "*.pdf")])
        if files:
            self.selected_paths.extend(files)
            count = len(collect_pdf_files(self.selected_paths))
            self.status_var.set(f"已选择 PDF：{count} 个")
            self.log(f"已选择 PDF 文件，本次共 {count} 个。")

    def choose_folder(self):
        folder = filedialog.askdirectory(title="选择发票文件夹")
        if folder:
            self.selected_paths.append(folder)
            count = len(collect_pdf_files(self.selected_paths))
            self.status_var.set(f"已选择文件夹，找到 PDF：{count} 个")
            self.log(f"已选择文件夹：{folder}")
            self.log(f"递归找到 PDF：{count} 个。")

    def choose_output(self):
        path = filedialog.asksaveasfilename(
            title="选择输出 Excel",
            defaultextension=".xlsx",
            initialfile=DEFAULT_EXCEL_NAME,
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if path:
            self.output_path_var.set(path)

    def start(self):
        pdf_files = collect_pdf_files(self.selected_paths)
        if not pdf_files:
            messagebox.showerror("没有 PDF", "请先选择 PDF 文件或包含 PDF 的文件夹。")
            return
        excel_path = Path(self.output_path_var.get().strip() or default_excel_path())
        self.progress["value"] = 0
        self.count_var.set("成功：0    失败：0")
        self.log(f"开始识别 {len(pdf_files)} 个 PDF。")
        self.log(f"输出 Excel：{excel_path}")
        self.worker = threading.Thread(target=self.run_worker, args=(pdf_files, excel_path), daemon=True)
        self.worker.start()

    def run_worker(self, pdf_files, excel_path: Path):
        success_rows = []
        success = 0
        failed = 0
        error_log = excel_path.parent / "error_log.txt"
        debug_log = excel_path.parent / "invoice_debug_log.txt"
        for index, pdf_path in enumerate(pdf_files, start=1):
            try:
                self.queue.put(("log", f"正在识别：{pdf_path.name}"))
                row = parse_invoice(pdf_path, debug_log)
                success_rows.append(row)
                success += 1
                self.queue.put(("log", f"识别成功：{pdf_path.name}"))
            except Exception as exc:
                failed += 1
                reason = str(exc)
                write_log(error_log, f"{pdf_path.name} - {reason}")
                self.queue.put(("log", f"识别失败：{pdf_path.name}，原因：{reason}"))
            self.queue.put(("progress", index, len(pdf_files), success, failed))
        try:
            if success_rows:
                write_excel(excel_path, success_rows)
            self.queue.put(("done", success, failed, str(excel_path)))
        except Exception as exc:
            write_log(error_log, "写入 Excel 失败：" + str(exc))
            self.queue.put(("log", "写入 Excel 失败：" + str(exc)))
            self.queue.put(("done", success, failed + len(success_rows), str(excel_path)))

    def poll_queue(self):
        try:
            while True:
                event = self.queue.get_nowait()
                if event[0] == "log":
                    self.log(event[1])
                elif event[0] == "progress":
                    _, current, total, success, failed = event
                    self.progress["value"] = current / total * 100
                    self.count_var.set(f"成功：{success}    失败：{failed}")
                elif event[0] == "done":
                    _, success, failed, excel_path = event
                    self.progress["value"] = 100
                    self.status_var.set("识别完成")
                    self.log(f"处理完成。成功 {success}，失败 {failed}。")
                    self.log(f"Excel 输出：{excel_path}")
                    messagebox.showinfo("完成", f"处理完成：成功 {success}，失败 {failed}")
        except Empty:
            pass
        self.root.after(200, self.poll_queue)

    def open_excel(self):
        path = Path(self.output_path_var.get().strip() or default_excel_path())
        if not path.exists():
            messagebox.showerror("文件不存在", "输出 Excel 还没有生成。")
            return
        os.startfile(path)

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    InvoiceApp().run()
